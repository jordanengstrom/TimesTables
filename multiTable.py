#! python3

# Imports modules and loads workbook
import openpyxl
from openpyxl.styles import Font

print('Opening workbook...')
wb = openpyxl.Workbook()
sheet = wb['Sheet']

# Ask for user input and convert to int
n = abs(int(input("Choose an integer greater than zero: ")))
fontObj1 = Font(bold=True)

# Populate the cells row by row
# Algorithm: 1i, 2i, 3i, ... , i*i
for i in range(2, n + 2):  # Leave the top left cell 'A1' blank
    sheet.cell(row=1, column=i).value = i - 1  # Top row
    sheet.cell(row=1, column=i).font = fontObj1

    sheet.cell(row=i, column=1).value = i - 1  # Left row
    sheet.cell(row=i, column=1).font = fontObj1

    for j in range(2, n + 2):
        sheet.cell(row=i, column=j).value = (i - 1) * (j - 1)


wb.save('multiTable.xlsx')



